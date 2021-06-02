from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.options import Options
import time
from datetime import datetime, date
from openpyxl import Workbook

#
# variables
#
categories = []

#
# methods
#

url = "http://www.examplesite.com"
username = ""
password = ""

def login(): # login funcion
    # access url
    driver.get(url)
    # input fields
    loginForm = '//*[@id="authorization-trigger"]'
    username_input = '//*[@id="cnpj"]'
    password_input = '//*[@id="pass"]'
    login_btn = '//*[@id="send2"]'
    # browser login actions
    driver.find_element_by_xpath(loginForm).click()
    driver.find_element_by_xpath(username_input).send_keys(username)
    driver.find_element_by_xpath(password_input).send_keys(password)
    driver.find_element_by_xpath(login_btn).click()

def getCategories():
    categoriesEl = driver.find_elements_by_css_selector('li.item.level0.level-top.parent')
    for category in categoriesEl:
        title = category.find_element_by_css_selector('span')
        url = category.find_element_by_css_selector('.menu-link')
        categories.append({ 'title': title.text, 'url': url.get_attribute('href')
    })

def loadCategory():

    result = []

    for category in categories:
        # if category['title'] != 'Energia Solar':
        if category['title'] == 'Ferramentas':
            # set pages loop
            driver.get(category['url']+'?p=1&product_list_limit=64')
            pagesCountEl = driver.find_element_by_css_selector('#toolbar-amount span:last-child')
            pagesCount = round(int(pagesCountEl.text) / 64)
            # collect data from pages
            for page in range(pagesCount):
                print('lendo categoria: ', category['url'] , 'página', page+1)
                driver.get(category['url']+'?p='+str(page+1)+'&product_list_limit=64')
                # result.append(*getProducts(category['title']))
                result += getProducts(category['title'])

    # create a file
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'name'
    ws['B1'] = 'price'
    ws['C1'] = 'stock'
    ws['D1'] = 'category'

    for i,res in enumerate(result, start=2):
        ws['A'+str(i)] = res['title']
        ws['B'+str(i)] = res['price']
        ws['C'+str(i)] = res['stock']
        ws['D'+str(i)] = res['category']

    # Save the file
    log = str(datetime.now())
    wb.save(log+".xlsx")

def getProducts(category):

    productsEl = driver.find_elements_by_css_selector('.product-item')

    products = []
    for product in productsEl:
        title = product.find_element_by_css_selector('.product-item-name')
        try:
            price = product.find_element_by_css_selector('[data-price-type="finalPrice"] .price')
            price.get_attribute('style') == "display: none;"

            products.append({"category": category, "title": title.text, "price": price.text, "stock": "true"})

        except NoSuchElementException:
            products.append({"category": category, "title": title.text, "price": "", "stock": "false"})

    return products


print('### starting script ###') # call methods
options = Options()
options.headless = True
# change this to your chromedriver location
driver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver', options=options)
time.sleep(1)
login()
getCategories()
loadCategory()
driver.close()
print('###      DONE!      ###') # done